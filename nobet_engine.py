import calendar
from datetime import date, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Font
from collections import defaultdict

monthly_stats = defaultdict(lambda: defaultdict(lambda: {"bayram":0,"haftasonu":0,"normal":0}))

MAX_SAME_WEEKDAY = 2
WEEKDAY_PENALTY = 1.5
DENGE_KATSAYI = 0.7
AGIR_GUN_FRENI = 0.4
MIN_GAP_DAYS = 14

eklenme_tarihi = {}
cikma_tarihi = {}
eklenme = {}

# =====================================
# GEÇMİŞ YÜK (SENİN VERİN AYNI KALACAK)
# =====================================

GECMIS_YUK = {
"YÖRÜKSELİM": {"bayram": 2, "haftasonu": 0, "normal": 2},
"CEYLAN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"KAZANCI": {"bayram": 2, "haftasonu": 0, "normal": 2},
"NESİBE": {"bayram": 3, "haftasonu": 1, "normal": 1},
"ONUR": {"bayram": 2, "haftasonu": 1, "normal": 2},
"DEMİRCİLER": {"bayram": 2, "haftasonu": 0, "normal": 2},
"ÇINAR": {"bayram": 2, "haftasonu": 1, "normal": 2},
"AKSU": {"bayram": 3, "haftasonu": 0, "normal": 2},
"AKPINAR": {"bayram": 2, "haftasonu": 0, "normal": 3},
"ÖZCAN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"YERHAN": {"bayram": 2, "haftasonu": 1, "normal": 0},
"MAĞRALI": {"bayram": 2, "haftasonu": 0, "normal": 3},
"TUNA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"VERESELİ PELİN": {"bayram": 1.9, "haftasonu": 1, "normal": 2},
"BAYRAM": {"bayram": 2, "haftasonu": 1, "normal": 2},
"MEHTAP": {"bayram": 2, "haftasonu": 0, "normal": 3},
"LAVANTA": {"bayram": 2, "haftasonu": 1, "normal": 1},
"NEFES": {"bayram": 2, "haftasonu": 1, "normal": 2},
"AKSÜT": {"bayram": 2, "haftasonu": 1, "normal": 2},
"NİMET": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SAĞOCAK": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ARSLANTÜRK": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ARZU": {"bayram": 2, "haftasonu": 1, "normal": 2},
"HARUN": {"bayram": 2, "haftasonu": 1, "normal": 1},
"VİLDAN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"KARAMANLI": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ORTASEKİ": {"bayram": 2, "haftasonu": 1, "normal": 2},
"EMİR": {"bayram": 2, "haftasonu": 1, "normal": 2},
"AYŞE": {"bayram": 2, "haftasonu": 1, "normal": 2},
"DOĞAN": {"bayram": 2, "haftasonu": 0, "normal": 2},
"REMZİ": {"bayram": 2, "haftasonu": 0, "normal": 1},
"SÜLEYMAN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SARIKAYA": {"bayram": 2, "haftasonu": 1, "normal": 1},
"HÜDAYİOĞLU": {"bayram": 3, "haftasonu": 0, "normal": 2},
"KARŞIYAKA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"KURTULUŞ": {"bayram": 3, "haftasonu": 1, "normal": 1},
"FİLİZ": {"bayram": 2, "haftasonu": 1, "normal": 2},
"İNCEER": {"bayram": 2, "haftasonu": 1, "normal": 2},
"CANPOLAT": {"bayram": 2, "haftasonu": 0, "normal": 3},
"VURAL": {"bayram": 2, "haftasonu": 1, "normal": 2},
"DAVARCIOĞLU": {"bayram": 2, "haftasonu": 1, "normal": 2},
"CEREN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SİMYA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"MURAT": {"bayram": 3, "haftasonu": 1, "normal": 2},
"HÜRRİYET": {"bayram": 3, "haftasonu": 1, "normal": 1},
"DEMET": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ÇOLAKOĞLU": {"bayram": 2, "haftasonu": 1, "normal": 2},
"YAĞMUR": {"bayram": 2, "haftasonu": 1, "normal": 2},
"KAYTAN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"AYSUN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"KOZANOĞLU": {"bayram": 3.2, "haftasonu": 0, "normal": 2},
"ÇİĞDEM": {"bayram": 2, "haftasonu": 1, "normal": 2},
"DERYA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"BAHADIR": {"bayram": 2, "haftasonu": 0, "normal": 4},
"NİSAN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"IHLAMUR": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SELİN": {"bayram": 2, "haftasonu": 1, "normal": 1},
"HÜSNA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"BİLGE": {"bayram": 2, "haftasonu": 1, "normal": 2},
"DORUK": {"bayram": 2, "haftasonu": 0, "normal": 3},
"BİNEVLER": {"bayram": 3.2, "haftasonu": 1, "normal": 1},
"DOĞA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"NEŞE SAYIT": {"bayram": 2, "haftasonu": 1, "normal": 2},
"MERT": {"bayram": 2, "haftasonu": 0, "normal": 2},
"ANNEM": {"bayram": 2, "haftasonu": 1, "normal": 2},
"MEHPARE": {"bayram": 2, "haftasonu": 1, "normal": 2},
"FURKAN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ONİKİŞUBAT": {"bayram": 2, "haftasonu": 1, "normal": 2},
"DÖKÜCÜ": {"bayram": 2, "haftasonu": 1, "normal": 1},
"ZÜMRA": {"bayram": 2, "haftasonu": 1, "normal": 1},
"ANADOLU": {"bayram": 3, "haftasonu": 1, "normal": 2},
"ÜNGÜT": {"bayram": 2, "haftasonu": 1, "normal": 1},
"GÜL": {"bayram": 3, "haftasonu": 0, "normal": 2},
"İRŞAD": {"bayram": 2, "haftasonu": 0, "normal": 3},
"GEMCİ": {"bayram": 2, "haftasonu": 0, "normal": 2},
"CANSU": {"bayram": 2, "haftasonu": 1, "normal": 1},
"ŞAHBAZ": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SAADET": {"bayram": 2, "haftasonu": 1, "normal": 2},
"NİŞANTAŞI": {"bayram": 2, "haftasonu": 1, "normal": 1},
"ŞENEL": {"bayram": 2, "haftasonu": 0, "normal": 3},
"EDA": {"bayram": 2, "haftasonu": 1, "normal": 1},
"ELVİN": {"bayram": 2, "haftasonu": 1, "normal": 1},
"FARAH HATİPOĞLU": {"bayram": 2, "haftasonu": 1, "normal": 1},
"BOĞAZİÇİ": {"bayram": 2, "haftasonu": 1, "normal": 2},
"OKAN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"KAYNAR": {"bayram": 2, "haftasonu": 0, "normal": 2},
"TEKEREK": {"bayram": 2, "haftasonu": 1, "normal": 2},
"MERVE": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ARISOY": {"bayram": 2, "haftasonu": 0, "normal": 3},
"NAR": {"bayram": 2, "haftasonu": 1, "normal": 2},
"LİMON": {"bayram": 3, "haftasonu": 0, "normal": 2},
"BOLAT": {"bayram": 2, "haftasonu": 0, "normal": 2},
"AKKÜNCÜ": {"bayram": 2, "haftasonu": 0, "normal": 2},
"BATUHAN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"GÜNEY": {"bayram": 2, "haftasonu": 0, "normal": 3},
"BÜŞRA ATA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"AKASYA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"GÖKTUĞ": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SU": {"bayram": 3, "haftasonu": 1, "normal": 1},
"HİLAL": {"bayram": 3, "haftasonu": 1, "normal": 2},
"ZEYNEP": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ESRA BÜYÜKDERELİ": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SERPİL": {"bayram": 3, "haftasonu": 1, "normal": 0},
"CENNET": {"bayram": 3, "haftasonu": 0, "normal": 3},
"KEREM": {"bayram": 2, "haftasonu": 1, "normal": 1},
"YATILI BÖLGE": {"bayram": 2, "haftasonu": 0, "normal": 3},
"VEZİR": {"bayram": 2, "haftasonu": 0, "normal": 3},
"HACETTEPE": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ERSOY": {"bayram": 2, "haftasonu": 1, "normal": 2},
"GÜLERYÜZ": {"bayram": 3, "haftasonu": 1, "normal": 1},
"ELMAS": {"bayram": 2, "haftasonu": 0, "normal": 2},
"GÖKÇE": {"bayram": 2, "haftasonu": 0, "normal": 3},
"CAN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"TEKİNŞEN": {"bayram": 3, "haftasonu": 1, "normal": 1},
"KOÇAK": {"bayram": 3, "haftasonu": 0, "normal": 1},
"SIHHAT": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ÇARE": {"bayram": 2, "haftasonu": 1, "normal": 2},
"PAKSOY": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ÇAĞATAY": {"bayram": 2, "haftasonu": 1, "normal": 2},
"BAL": {"bayram": 2, "haftasonu": 1, "normal": 2},
"BİLAL": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ESRA AKSOY": {"bayram": 2, "haftasonu": 1, "normal": 2},
"AVŞAROĞLU": {"bayram": 3, "haftasonu": 1, "normal": 2},
"RAİKA DOKUYUCU": {"bayram": 2, "haftasonu": 1, "normal": 2},
"NATUREL": {"bayram": 2, "haftasonu": 2, "normal": 1},
"FLORA": {"bayram": 3, "haftasonu": 1, "normal": 2},
"İBNİ SİNA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"KEVSER": {"bayram": 3.2, "haftasonu": 1, "normal": 2},
"ARAS": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SOLMAZ": {"bayram": 3, "haftasonu": 1, "normal": 1},
"SÜMEN": {"bayram": 2, "haftasonu": 1, "normal": 3},
"OCAK": {"bayram": 2, "haftasonu": 1, "normal": 3},
"YALÇIN": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ESRA": {"bayram": 2, "haftasonu": 0, "normal": 3},
"ASLANBEY": {"bayram": 2, "haftasonu": 1, "normal": 3},
"NİL": {"bayram": 2, "haftasonu": 1, "normal": 3},
"GÜVEN": {"bayram": 2, "haftasonu": 0, "normal": 3},
"KILIÇ": {"bayram": 3, "haftasonu": 1, "normal": 2},
"ELİF": {"bayram": 2, "haftasonu": 1, "normal": 2},
"YEDİTEPE": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ERDİ": {"bayram": 2, "haftasonu": 1, "normal": 3},
"PİRİ REİS 1453": {"bayram": 2, "haftasonu": 1, "normal": 2},
"DEFNE": {"bayram": 3.2, "haftasonu": 0, "normal": 4},
"YUNUS EMRE": {"bayram": 3.2, "haftasonu": 1, "normal": 2},
"LOKMAN": {"bayram": 3, "haftasonu": 1, "normal": 2},
"TUĞRUL": {"bayram": 2, "haftasonu": 1, "normal": 2},
"BESLER": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SEMA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SIDIKA": {"bayram": 2, "haftasonu": 1, "normal": 3},
"GAMZE": {"bayram": 3, "haftasonu": 1, "normal": 1},
"KÜMBET": {"bayram": 2, "haftasonu": 0, "normal": 3},
"SERKAN": {"bayram": 3, "haftasonu": 1, "normal": 0},
"ŞİFA": {"bayram": 3, "haftasonu": 1, "normal": 2},
"DEVA": {"bayram": 3, "haftasonu": 1, "normal": 1},
"CEM": {"bayram": 2, "haftasonu": 1, "normal": 3},
"MAVİ": {"bayram": 2, "haftasonu": 1, "normal": 3},
"TUĞBA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"LEYLA DOKUMACI": {"bayram": 1.9, "haftasonu": 1, "normal": 2},
"RAMAZANOĞLU": {"bayram": 2, "haftasonu": 1, "normal": 2},
"DENİZ": {"bayram": 2, "haftasonu": 1, "normal": 3},
"AYLİN TATLI": {"bayram": 2, "haftasonu": 1, "normal": 3},
"NECİP FAZIL": {"bayram": 2, "haftasonu": 1, "normal": 3},
"ÖZDEMİR": {"bayram": 2, "haftasonu": 1, "normal": 3},
"YÜCEL": {"bayram": 1.9, "haftasonu": 1, "normal": 2},
"YILDIRIM": {"bayram": 2, "haftasonu": 1, "normal": 3},
"TOMAR": {"bayram": 2, "haftasonu": 1, "normal": 3},
"GÜNEŞ": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ŞİMŞEK": {"bayram": 2, "haftasonu": 1, "normal": 2},
"KARAKÜÇÜK": {"bayram": 2, "haftasonu": 1, "normal": 3},
"POYRAZ": {"bayram": 2, "haftasonu": 1, "normal": 3},
"DERMAN": {"bayram": 2, "haftasonu": 1, "normal": 3},
"KARACAOĞLAN": {"bayram": 3, "haftasonu": 1, "normal": 3},
"ESMA": {"bayram": 2, "haftasonu": 1, "normal": 3},
"AYŞEGÜL": {"bayram": 2, "haftasonu": 1, "normal": 2},
"ÖZLEM": {"bayram": 2, "haftasonu": 1, "normal": 3},
"ALYA": {"bayram": 2, "haftasonu": 1, "normal": 2},
"SEZAL": {"bayram": 3.2, "haftasonu": 1, "normal": 2},
"ÇEVİK": {"bayram": 2, "haftasonu": 1, "normal": 2},
"GAZİ": {"bayram": 1.8, "haftasonu": 0, "normal": 2},
"GLSAH": {"bayram": 1.8, "haftasonu": 1, "normal": 2},
"CEYDA İLHAN": {"bayram": 1.8, "haftasonu": 1, "normal": 2},
"İBNİ SİNA 2": {"bayram": 1.8, "haftasonu": 0, "normal": 3}

}

# =====================================
# TATİL FONKSİYONLARI
# =====================================

def turkiye_tatilleri(year):
    return {
        date(year,1,1),
        date(year,4,23),
        date(year,5,1),
        date(year,5,19),
        date(year,7,15),
        date(year,8,30),
        date(year,10,29),
        date(2026,3,20),
        date(2026,3,21),
        date(2026,3,22),
        date(2026,5,27),
        date(2026,5,28),
        date(2026,5,29),
        date(2026,5,30),
    }

def arefe_gunleri(year):
    return {
        date(2026,3,19),
        date(2026,5,26)
    }

# =====================================
# GÜN KATSAYISI
# =====================================

def day_weight(d,tatil,arefe):

    if d in tatil or d.weekday()==6:
        return 2.0

    if d.weekday()==5 or d in arefe:
        return 1.5

    return 1.0

# =====================================
# SKOR HESABI
# =====================================

def score_person(p,d,w,totals,counts,weekday_stats,last_dates):

    skor = totals[p]*DENGE_KATSAYI + counts[p]

    if weekday_stats[p][d.weekday()] >= MAX_SAME_WEEKDAY:
        skor += WEEKDAY_PENALTY

    if p in last_dates:
        gap = (d-last_dates[p]).days
        skor -= min(gap,30)*0.05

    if w>1.4:
        skor += AGIR_GUN_FRENI * weekday_stats[p][d.weekday()]

    return skor + random.random()*0.01

# =====================================
# ECZANE SEÇİMİ
# =====================================

def zorunlu_secim(grup,d,w,tatil,totals,counts,weekday_stats,last_dates,bayram_stats):

    kademe1=[]
    kademe2=[]
    kademe3=[]

    for p in grup:

        if p in eklenme_tarihi and d < eklenme_tarihi[p]:
            continue

        if p in cikma_tarihi and d >= cikma_tarihi[p]:
            continue

        gap = (d-last_dates[p]).days if p in last_dates else 999

        if gap >= MIN_GAP_DAYS and weekday_stats[p][d.weekday()] < MAX_SAME_WEEKDAY:
            kademe1.append(p)

        elif weekday_stats[p][d.weekday()] < MAX_SAME_WEEKDAY:
            kademe2.append(p)

        else:
            kademe3.append(p)

    adaylar = kademe1 or kademe2 or kademe3 or grup

    if d in tatil:
        min_b = min(bayram_stats[p] for p in adaylar)
        adaylar = [p for p in adaylar if bayram_stats[p]==min_b]

    return min(adaylar,key=lambda p: score_person(p,d,w,totals,counts,weekday_stats,last_dates))

# =====================================
# GRUPLAR (SENİN LİSTEN AYNI)
# =====================================

def create_groups():

    groups = {
        "A1": ["ŞAHBAZ","BATUHAN","İRŞAD","MEHPARE","GÜL","GEMCİ","RAİKA DOKUYUCU","ANADOLU","CANSU"],
        "A2": ["GÖKTUĞ","GÜNEY","NEŞE SAYIT","FLORA","BOĞAZİÇİ","HÜSNA","GÜLERYÜZ","ÜNGÜT","LİMON"],
        "A3": ["ESRA AKSOY","NAR","AVŞAROĞLU","MERT","GAZİ","DÖKÜCÜ","AKKÜNCÜ","ANNEM","BİLAL"],

        "B1": 
["NİŞANTAŞI","EDA","ŞENEL","FARAH HATİPOĞLU","MERVE","ELVİN","BİNEVLER","TEKEREK"],
        "B2":
["SERPİL","FURKAN","ARISOY","SU","AKASYA","CENNET","ÇAĞATAY","İNCEER","HİLAL"],       
        "B3":
["KAYNAR","NATUREL","OKAN","ESRA BÜYÜKDERELİ","BÜŞRA ATA","BAL","GLSAH","CEYDA İLHAN"],
        "C1":
["FİLİZ","ÇİĞDEM","KARŞIYAKA","DEMET","MURAT","HÜDAYİOĞLU","CANPOLAT", "DAVARCIOĞLU"],
        "C2": 
["PAKSOY","ZEYNEP","GÖKÇE","HÜRRİYET","KURTULUŞ","KAYTAN","KOZANOĞLU","İBNİ SİNA 2","BİLGE"],
        "C3": 
["NİSAN","SELİN","IHLAMUR","DORUK","ELMAS","ONİKİŞUBAT","BOLAT","ZÜMRA","SAADET"], 

        "D1": ["YAĞMUR","AYSUN","CEREN","BAHADIR","SİMYA","DERYA","VURAL","ÇOLAKOĞLU"],
        "D2": ["ÇARE","KOÇAK","KEREM","CAN","TEKİNŞEN","SIHHAT","YATILI BÖLGE","VEZİR","DOĞA"],

        "D3": ["ERSOY","HACETTEPE","SAĞOCAK","ARZU","DEMİRCİLER","ORTASEKİ","VİLDAN","KARAMANLI","SARIKAYA"],
        

        "E1": 
["ONUR","ÖZCAN","AKSU","VERESELİ PELİN","NİMET","AKPINAR","ARSLANTÜRK","SÜLEYMAN","NESİBE"],
        "E2": ["KAZANCI","DOĞAN","HARUN","EMİR","LAVANTA","YERHAN","REMZİ","MAĞRALI","YÖRÜKSELİM"],
        "E3": ["CEYLAN","ÇINAR","TUNA","AYŞE","AKSÜT","NEFES","MEHTAP","BAYRAM"],

        "F1": ["ARAS","YUNUS EMRE","ERDİ","YEDİTEPE","DEFNE","ELİF","OCAK","KEVSER","DENİZ"],
        "F2": ["SÜMEN","TUĞBA","MAVİ","GAMZE","YALÇIN","SOLMAZ","KÜMBET","SERKAN","İBNİ SİNA"],
        "F3": 
["ÖZLEM","KARAKÜÇÜK","AYLİN TATLI","ÖZDEMİR","NECİP FAZIL","CEM","RAMAZANOĞLU","KILIÇ","LEYLA DOKUMACI"],

        "G1": 
["SIDIKA","TUĞRUL","BESLER","SEMA","NİL","ASLANBEY", "ESRA","GÜVEN"],
        "G2": ["YILDIRIM","DEVA","ŞİFA","SEZAL","TOMAR","YÜCEL","LOKMAN","ŞİMŞEK"],
        "G3": ["GÜNEŞ","KARACAOĞLAN","ÇEVİK","PİRİ REİS 1453","DERMAN","ESMA","AYŞEGÜL","POYRAZ","ALYA"],

    }

    # STREAMLIT EKLENEN ECZANE

    for eczane,data in eklenme.items():

        grup = data["grup"]

        if grup in groups:

            if eczane not in groups[grup]:

                groups[grup].append(eczane)

    return groups

# =====================================
# ROTASYON
# =====================================

KOMB_ABC=[("A1","B2","C3"),
("B1","C2","A3"),
("C1","A2","B3"),
("A1","C2","B3"),
("B1","A2","C3"),
("C1","B2","A3")]

KOMB_DEG=[("D1","E2","G3"),
("E1","G2","D3"),
("G1","D2","E3"),
("D1","G2","E3"),
("E1","D2","G3"),
("G1","E2","D3")]
F_ROTASYON=["F1","F2","F3"]

# =====================================
# AYLIK PLAN
# =====================================

def generate_month(groups,year,month,totals,counts,weekday_stats,bayram_stats,last_dates):

    tatil = turkiye_tatilleri(year)
    arefe = arefe_gunleri(year)

    first=date(year,month,1)
    dim=calendar.monthrange(year,month)[1]

    schedule={}

    for i in range(dim):

        d=first+timedelta(days=i)
        w=day_weight(d,tatil,arefe)

        picks={}

        for g in KOMB_ABC[i%6] + KOMB_DEG[i%6]:

            pick = zorunlu_secim(
                groups[g],d,w,tatil,
                totals,counts,weekday_stats,last_dates,bayram_stats
            )

            picks[g]=pick

            totals[pick]+=w
            counts[pick]+=1
            weekday_stats[pick][d.weekday()]+=1
            last_dates[pick]=d

            if d in tatil:
                bayram_stats[pick]+=1

            key=(d.year,d.month)

            if d in tatil:
                monthly_stats[pick][key]["bayram"]+=1
            elif d.weekday()>=5:
                monthly_stats[pick][key]["haftasonu"]+=1
            else:
                monthly_stats[pick][key]["normal"]+=1

        fg=F_ROTASYON[i%3]

        pick = zorunlu_secim(
            groups[fg],d,w,tatil,
            totals,counts,weekday_stats,last_dates,bayram_stats
        )

        picks[fg]=pick

        totals[pick]+=w
        counts[pick]+=1
        weekday_stats[pick][d.weekday()]+=1
        last_dates[pick]=d
        key=(d.year,d.month)
        if d in tatil:
            monthly_stats[pick][key]["bayram"]+=1
        elif d.weekday()>=5:
            monthly_stats[pick][key]["haftasonu"]+=1
        else:
            monthly_stats[pick][key]["normal"]+=1

        schedule[d]=picks

    return schedule

# =====================================
# MAIN
# =====================================
def main(y,m,nm):

    random.seed(42)
    groups = create_groups()

    totals = {p:0 for g in groups.values() for p in g}
    counts = {p:0 for g in groups.values() for p in g}
    weekday_stats = {p:{i:0 for i in range(7)} for p in totals}
    bayram_stats = {p:0 for p in totals}
    for p,v in GECMIS_YUK.items():

        if p not in totals:
            continue

        kats = v["normal"] + v["haftasonu"]*1.5 + v["bayram"]*2

        totals[p] += kats
        counts[p] += v["normal"] + v["haftasonu"] + v["bayram"]
        bayram_stats[p] += v["bayram"]

        weekday_stats[p][5] += v["haftasonu"]//2
        weekday_stats[p][6] += v["haftasonu"]//2
        
    last_dates = {}

    wb = Workbook()

    gun = ["Pzt","Salı","Çarş","Perş","Cuma","Ctesi","Pazar"]
    header = ["Tarih","Gün"] + list(groups.keys())

    for k in range(nm):

        year = y + ((m-1+k)//12)
        month = ((m-1+k)%12)+1

        ws = wb.create_sheet(f"{year}-{month:02d}")
        ws.append(header)

        sched = generate_month(
            groups,
            year,
            month,
            totals,
            counts,
            weekday_stats,
            bayram_stats,
            last_dates
        )

        for d,p in sorted(sched.items()):

            row = [
                d.strftime("%d.%m.%Y"),
                gun[d.weekday()]
            ]

            for g in groups:
                row.append(p.get(g,""))

            ws.append(row)

        for c in ws[1]:
            c.font = Font(bold=True)

    # ==============================
    # GENEL OZET SAYFASI
    # ==============================
        # ==============================
    # GENEL OZET SAYFASI
    # ==============================

    summary = wb.create_sheet("GENEL OZET")

    summary.append([
        "Eczane",
        "Grup",
        "Geçmiş Katsayı",
        "Geçmiş Bayram",
        "Toplam Nöbet",
        "Toplam Katsayı",
        "Bayram",
        "Pzt","Salı","Çarş","Perş","Cuma","Ctesi","Pazar"
    ])

    eczane_grup = {p:g for g,plist in groups.items() for p in plist}

    for p in totals:

        gecmis = GECMIS_YUK.get(p,{"bayram":0,"haftasonu":0,"normal":0})

        gecmis_katsayi = (
            gecmis["bayram"]*2 +
            gecmis["haftasonu"]*1.5 +
            gecmis["normal"]
        )

        toplam_katsayi = gecmis_katsayi + totals[p]
        toplam_bayram = gecmis["bayram"] + bayram_stats[p]

        summary.append([
            p,
            eczane_grup.get(p,""),
            round(gecmis_katsayi,2),
            gecmis["bayram"],
            counts[p],     
            round(toplam_katsayi,2),
            bayram_stats[p],
            weekday_stats[p][0],
            weekday_stats[p][1],
            weekday_stats[p][2],
            weekday_stats[p][3],
            weekday_stats[p][4],
            weekday_stats[p][5],
            weekday_stats[p][6]
        ])

    for c in summary[1]:
        c.font = Font(bold=True)

    wb.remove(wb["Sheet"])

    wb.save("Son.xlsx")
   
    # ==============================
    # AYLIK DETAY EXCEL
    # ==============================

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "AYLIK DETAY"

    ws2.append([
        "Eczane",
        "Yıl",
        "Ay",
        "Bayram",
        "Hafta Sonu",
        "Normal"
    ])

        for eczane in sorted(totals.keys()):

        for (yil,ay),veri in sorted(monthly_stats[eczane].items()):

            ws2.append([
                eczane,
                yil,
                ay,
                veri["bayram"],
                veri["haftasonu"],
                veri["normal"]
            ])

    for c in ws2[1]:
        c.font = Font(bold=True)

    wb2.save("aylik_nobet_data.xlsx")

    return "Son.xlsx","aylik_nobet_data.xlsx"


# =====================================
# STREAMLIT ÇAĞIRMA
# =====================================

def run_schedule(y,m,nm,eklenme_input={},cikma_input={}):

    global eklenme_tarihi
    global cikma_tarihi
    global eklenme

    eklenme = eklenme_input

    eklenme_tarihi = {k:v["tarih"] for k,v in eklenme_input.items()}
    cikma_tarihi = cikma_input

    return main(y,m,nm)
